#! Python3
# test.py

# -*- coding: utf-8 -*-
from openpyxl import Workbook

#-----------Create a Workbook----------|

# Create new Workbook (spreadsheet)
wb = Workbook()

# Select active Sheet
dailyfocus = wb.active

# Create new title
dailyfocus.title = "Daily Focus"

# Change color of Workbook tab (at the bottom of the sheet)
dailyfocus.sheet_properties.tabColor = "2abdfc"  # Hex code / rrggbb code

# Create a new sheet
dailydelegation = wb.create_sheet("Daily Delegation")
dailydelegation.sheet_properties.tabColor = "fcbd00"

# Review names of all worksheets of workbook
print(wb.sheetnames)

# Assign a worksheet as a variable calling Workbook object
worksheet = wb["Daily Delegation"]

# Create copy of a worksheet
target = wb.copy_worksheet(dailyfocus)

#===========Playing with Data============|

#-----------Accessing one cell-----------|

# Access a cell
dailyfocus_a1 = dailyfocus['A1']

# Directly assign a value.
dailyfocus['A1'] = "Date"

# Access cells using row and column notation
b2 = dailyfocus.cell(row=1, column=2, value='Today')

#-----------Accessing many cells------------|

# Access a range of cells using slicing
cell_range = dailyfocus['A2':'A5']

# Access column
c = dailyfocus['C']

# Access range of columns
cd = dailyfocus['C:D']

# Access row
row10 = dailyfocus[10]

# Access range of rows
row_range = dailyfocus[5:10]

# Iterate through rows using iter_rows method
iterated_rows = []

for row in dailyfocus.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        iterated_rows.append(cell)


# Iterate through columns using iter_cols method
iterated_columns = []

for col in dailyfocus.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        iterated_columns.append(cell)

# Iterate through all the rows or columns of a file
dailyfocus.rows  # Goes through all the rows
dailyfocus.columns  # Goes through all the columns

#----------Data Storage-----------|

# Assign a value to a cell
c.value = 'hello, world'
print(c.value)
d.value = 3.14
print(d.value)

#-----------Loading from a File----------|
from openpyxl import load_workbook
wb2 = load_workbook('test.xlsx')
print(wb2.sheetnames)

# Save the spreadsheet
wb.save('new.xlsx')

